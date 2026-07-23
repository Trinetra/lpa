"""Daily backup: a restorable mongodump archive + a human-readable Excel
workbook, zipped together and uploaded to a "Backups" folder in the
teacher's Google Drive (via the same OAuth connection used for Calendar).

Two formats in one ZIP because they serve different purposes:
- mongo.archive.gz: exact, complete, restorable with
  `mongorestore --archive=mongo.archive.gz --gzip`. This is the real
  disaster-recovery path.
- records.xlsx: human-readable, one sheet per record type, openable in
  Excel/Sheets without any tooling, for "let me just look something up"
  during an outage.
"""

import io
import logging
import os
import subprocess
import zipfile
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

from db import db
from services import calendar as calendar_service

logger = logging.getLogger(__name__)

BACKUP_FOLDER_NAME = "Lakshmi Studio Ledger — Backups"

# Collections that are actual user records worth backing up. Deliberately
# excludes internal/system collections: password_reset_tokens (short-lived,
# security-sensitive), invoice_counters (regenerable), and users itself is
# handled separately below with the password hash stripped.
_RECORD_COLLECTIONS = [
    "students", "classes", "payments", "invoices",
    "schedule_blocks",
    "tours", "tour_stops", "tour_expenses", "tour_invoices",
    "tour_contacts", "tour_todos", "tour_checkins",
]


def _flatten(doc: dict, prefix: str = "") -> dict:
    """Flatten nested dicts (e.g. invoice.summary.total_billed) into
    dotted-key columns so each sheet stays one-row-per-record instead of
    needing a second sheet for nested data."""
    out = {}
    for k, v in doc.items():
        key = f"{prefix}{k}"
        if isinstance(v, dict):
            out.update(_flatten(v, f"{key}."))
        elif isinstance(v, list):
            out[key] = "; ".join(str(x) for x in v)
        else:
            out[key] = v
    return out


async def _collection_rows(name: str, owner_id: str) -> list:
    docs = []
    async for d in db[name].find({"owner_id": owner_id}):
        d["_id"] = str(d["_id"])
        docs.append(_flatten(d))
    return docs


async def _profile_row(owner_id: str) -> dict:
    from bson import ObjectId
    user = await db.users.find_one({"_id": ObjectId(owner_id)})
    user = user or {}
    # Never export the password hash.
    user.pop("password_hash", None)
    user["_id"] = str(user.get("_id", ""))
    return _flatten(user)


def _sheet_title(name: str) -> str:
    # Excel sheet names cap at 31 chars and can't contain []:*?/\ — none of
    # our collection names hit either limit, but title-case for readability.
    return name.replace("_", " ").title()[:31]


def _write_sheet(wb, title: str, rows: list) -> None:
    ws = wb.create_sheet(title=title)
    if not rows:
        ws.append(["(no records)"])
        return
    fieldnames = []
    for row in rows:
        for k in row.keys():
            if k not in fieldnames:
                fieldnames.append(k)
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])
    for i, f in enumerate(fieldnames, start=1):
        width = max(len(f), *(len(str(row.get(f, ""))) for row in rows)) if rows else len(f)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)
    ws.freeze_panes = "A2"


def _run_mongodump() -> Optional[bytes]:
    """Run mongodump against the app's own database, return the gzipped
    archive bytes, or None if the mongodump binary/connection isn't
    available (backup still proceeds with records.xlsx only in that case)."""
    mongo_url = os.environ.get("MONGO_URL")
    db_name = os.environ.get("DB_NAME")
    if not mongo_url or not db_name:
        return None
    try:
        result = subprocess.run(
            ["mongodump", f"--uri={mongo_url}", f"--db={db_name}",
             "--archive", "--gzip"],
            capture_output=True, timeout=120, check=True,
        )
        return result.stdout
    except Exception as e:
        logger.error(f"mongodump failed: {e}")
        return None


async def build_backup_zip(owner_id: str) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest_lines = [
            "Lakshmi Studio Ledger — Backup",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            "",
            "Contents:",
            "  mongo.archive.gz — full database archive. To restore:",
            "    mongorestore --archive=mongo.archive.gz --gzip",
            "  records.xlsx — human-readable, one sheet per record type.",
            "",
        ]

        archive_bytes = _run_mongodump()
        if archive_bytes:
            zf.writestr("mongo.archive.gz", archive_bytes)
            manifest_lines.append("mongo.archive.gz: included")
        else:
            manifest_lines.append("mongo.archive.gz: NOT included (mongodump unavailable — records.xlsx only)")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # drop the default empty sheet

        for name in _RECORD_COLLECTIONS:
            rows = await _collection_rows(name, owner_id)
            if rows:
                _write_sheet(wb, _sheet_title(name), rows)
                manifest_lines.append(f"{_sheet_title(name)} sheet: included ({len(rows)} rows)")

        profile_row = await _profile_row(owner_id)
        _write_sheet(wb, "Studio Profile", [profile_row])
        manifest_lines.append("Studio Profile sheet: included (password not exported)")

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        zf.writestr("records.xlsx", xlsx_buf.getvalue())

        zf.writestr("MANIFEST.txt", "\n".join(manifest_lines))

    buf.seek(0)
    return buf.read()


def _find_or_create_backup_folder(drive_service) -> str:
    query = (
        f"name='{BACKUP_FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' "
        "and trashed=false"
    )
    results = drive_service.files().list(q=query, fields="files(id)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    folder = drive_service.files().create(
        body={"name": BACKUP_FOLDER_NAME, "mimeType": "application/vnd.google-apps.folder"},
        fields="id",
    ).execute()
    return folder["id"]


async def run_daily_backup(owner_id: str) -> dict:
    """Build the backup ZIP and upload it to the teacher's Drive. Returns a
    status dict; never raises — a failed backup shouldn't crash whatever
    triggered it (a cron job, a manual button click)."""
    from bson import ObjectId
    user = await db.users.find_one({"_id": ObjectId(owner_id)})
    if not user or not user.get("google_refresh_token"):
        return {"ok": False, "reason": "Google not connected"}

    try:
        # Explicitly request only the Drive scope here, not the shared
        # SCOPES list — an account that connected before Drive access was
        # added only holds calendar.CALENDAR_SCOPES on its stored token, so
        # requesting more than that fails the refresh outright. This will
        # surface as exactly that until she reconnects (which grants both).
        creds = await calendar_service.get_credentials(owner_id, scopes=calendar_service.DRIVE_SCOPES)
        if not creds:
            return {"ok": False, "reason": "Google not connected"}

        zip_bytes = await build_backup_zip(owner_id)
        drive_service = build("drive", "v3", credentials=creds)
        folder_id = _find_or_create_backup_folder(drive_service)

        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filename = f"backup_{date_str}.zip"
        media = MediaIoBaseUpload(io.BytesIO(zip_bytes), mimetype="application/zip", resumable=False)
        drive_service.files().create(
            body={"name": filename, "parents": [folder_id]},
            media_body=media,
            fields="id",
        ).execute()

        return {"ok": True, "filename": filename, "size": len(zip_bytes)}
    except Exception as e:
        logger.error(f"Daily backup failed: {e}")
        return {"ok": False, "reason": str(e)}
